#!/usr/bin/env python3
"""Parse BalVoi_30 Content.xlsx and scan pre-rendered audio folders → config JSON."""

from __future__ import annotations

import json

import openpyxl

from balvoi.paths import ROOT

XLSX = ROOT / "BalVoi_30 Content.xlsx"
CONFIG = ROOT / "config"
EDITIONS = CONFIG / "editions.json"

# xlsx column header → edition slug
SHEET_LANG_COLUMNS = [
    ("English Welcome", "en"),
    ("Spanish Welcome", "es"),
    ("Portuguese Welcome", "pt"),
    ("French Welcome", "fr"),
    ("German Welcome", "de"),
    ("Arabic Welcome", "ar"),
    ("Russian Welcome", "ru"),
    ("Turkish Welcome", "tr"),
]

SHEET_TO_KEY = {
    "Welcome": "welcome",
    "Started": "started",
    "Right Back": "right_back",
    "Ad 1": "ad_1",
    "Welcome Back": "welcome_back",
    "Ad 2": "ad_2",
    "Thank You": "thank_you",
}


def load_editions() -> list[dict]:
    return json.loads(EDITIONS.read_text(encoding="utf-8"))["editions"]


def parse_xlsx_segments() -> dict:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    segments: dict[str, dict[str, list[str]]] = {
        key: {slug: [] for _, slug in SHEET_LANG_COLUMNS}
        for key in SHEET_TO_KEY.values()
    }

    slug_by_header_prefix = {}
    for header, slug in SHEET_LANG_COLUMNS:
        # "English Welcome" → prefix "English"
        slug_by_header_prefix[header.split()[0]] = slug

    for sheet_name, seg_key in SHEET_TO_KEY.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_to_slug = {}
        for i, h in enumerate(headers):
            for prefix, slug in slug_by_header_prefix.items():
                if h.startswith(prefix):
                    col_to_slug[i] = slug
                    break

        for row in ws.iter_rows(min_row=2, values_only=True):
            for i, slug in col_to_slug.items():
                if i < len(row):
                    text = str(row[i] or "").strip()
                    if text:
                        segments[seg_key][slug].append(text)

    return segments


def find_mp3_assets(edition: dict) -> dict:
    folder = ROOT / edition["contentFolder"]
    assets: dict[str, list[str]] = {"ad_1": [], "ad_2": [], "right_back": []}

    mapping = {
        "ad_1": ["Ad 1"],
        "ad_2": ["Ad 2"],
        "right_back": ["Right Back", "Right Back "],
    }

    for key, subdirs in mapping.items():
        for sub in subdirs:
            sub_path = folder / sub
            if not sub_path.is_dir():
                continue
            for mp3 in sorted(sub_path.glob("*.mp3")):
                rel = mp3.relative_to(ROOT).as_posix()
                assets[key].append(rel)
        assets[key] = sorted(set(assets[key]))

    return assets


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing {XLSX}")

    editions = load_editions()
    segments = parse_xlsx_segments()

    assets_by_edition = {}
    for ed in editions:
        assets_by_edition[ed["id"]] = find_mp3_assets(ed)

    CONFIG.mkdir(parents=True, exist_ok=True)
    (CONFIG / "segments.json").write_text(
        json.dumps(segments, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (CONFIG / "assets.json").write_text(
        json.dumps(assets_by_edition, indent=2), encoding="utf-8"
    )

    missing = []
    for ed in editions:
        a = assets_by_edition[ed["id"]]
        for key in ("ad_1", "ad_2", "right_back"):
            if not a[key]:
                missing.append(f"{ed['id']}: {key}")

    print(f"Wrote {CONFIG / 'segments.json'}")
    print(f"Wrote {CONFIG / 'assets.json'}")
    if missing:
        print("WARN missing pre-rendered assets:")
        for m in missing:
            print(f"  - {m}")


if __name__ == "__main__":
    main()
